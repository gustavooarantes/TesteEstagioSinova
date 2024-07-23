from openpyxl import load_workbook

# Carrega o Excel workbook
file_path = 'TesteEstagio.xlsx'
workbook = load_workbook(filename=file_path)

sheet = workbook.active

# Define headers baseadas na estrutura
headers = ['DATA VENDA', 'CLIENTE', 'Preço Compra', 'Moeda', 'Preço Venda', 'Quantidade']

# Carrega os dados a partir da sétima linha
sales_data = []
for row in sheet.iter_rows(min_row=7, max_col=6, values_only=True):
    sales_data.append(dict(zip(headers, row)))

# Extrai as cotações
exchange_rates = {}
for row in sheet.iter_rows(min_row=8, min_col=17, max_col=18, values_only=True):
    if row[0] is not None and row[1] is not None:
        exchange_rates[row[0]] = row[1]

# Calcula o valor da margem e a margem percentual
def calculate_margins(data, exchange_rates):
    margins_percentual = []
    margins_value = []
    
    for entry in data:
        data_venda = entry['DATA VENDA']
        cliente = entry['CLIENTE']
        preco_compra = float(entry['Preço Compra'])
        moeda = entry['Moeda']
        preco_venda = float(entry['Preço Venda'])
        quantidade = float(entry['Quantidade'])
        
        exchange_rate = exchange_rates.get(data_venda, 1)
        
        # Converte de USD para BRL
        if moeda == 'USD':
            preco_compra_brl = preco_compra * exchange_rate
            preco_venda_brl = preco_venda * exchange_rate
        else:
            preco_compra_brl = preco_compra
            preco_venda_brl = preco_venda
        
        # Calcula as margens
        margin_percentual = (preco_venda - preco_compra) / preco_venda
        margin_value = (preco_venda_brl - preco_compra_brl) * quantidade
        
        margins_percentual.append((cliente, margin_percentual))
        margins_value.append((cliente, margin_value))
    
    return margins_percentual, margins_value

# Executa a função de calcular as margens
margins_percentual, margins_value = calculate_margins(sales_data, exchange_rates)

# Ordena as 5 maiores margens de forma decrescente
top_5_margins_percentual = sorted(margins_percentual, key=lambda x: x[1], reverse=True)[:5]
top_5_margins_value = sorted(margins_value, key=lambda x: x[1], reverse=True)[:5]

print("Top 5 clientes com maior margem percentual:")
for cliente, margem in top_5_margins_percentual:
    print(f"Cliente: {cliente}, Margem Percentual: {margem:.2%}")

print("\nTop 5 clientes com maior valor de margem em reais:")
for cliente, margem in top_5_margins_value:
    print(f"Cliente: {cliente}, Valor de Margem: R${margem:,.2f}")
